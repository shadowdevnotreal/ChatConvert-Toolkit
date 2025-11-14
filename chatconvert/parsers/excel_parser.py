"""
Excel Parser - Parse Excel spreadsheets containing chat logs.

Supports:
- XLSX files (Office 2007+)
- XLS files (Office 97-2003)

Expected format similar to CSV:
- First row: headers (timestamp, username, message)
- Subsequent rows: chat messages
- Supports multiple sheets
"""

from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from .base_parser import BaseParser
from ..models import Conversation, Message, Participant, MessageType


class ExcelParser(BaseParser):
    """Parse Excel format chat logs."""

    # Common column name variations (same as CSV)
    TIMESTAMP_COLUMNS = ['timestamp', 'time', 'date', 'datetime', 'sent_at', 'created_at']
    USERNAME_COLUMNS = ['username', 'user', 'sender', 'from', 'author', 'name']
    MESSAGE_COLUMNS = ['message', 'content', 'text', 'body']

    def can_parse(self, file_path: str) -> bool:
        """Check if file is Excel format."""
        ext = self._get_file_extension(file_path)

        if ext == 'xlsx' and HAS_OPENPYXL:
            return True
        elif ext == 'xls' and HAS_XLRD:
            return True

        return False

    def parse(self, file_path: str) -> Conversation:
        """
        Parse Excel file into Conversation object.

        Args:
            file_path: Path to Excel file

        Returns:
            Conversation object

        Raises:
            ValueError: If Excel format is invalid
            FileNotFoundError: If file doesn't exist
            ImportError: If required library not installed
        """
        self._validate_file(file_path)
        ext = self._get_file_extension(file_path)

        self.logger.info(f"Parsing Excel file: {file_path}")

        if ext == 'xlsx':
            return self._parse_xlsx(file_path)
        elif ext == 'xls':
            return self._parse_xls(file_path)
        else:
            raise ValueError(f"Unsupported Excel format: {ext}")

    def _parse_xlsx(self, file_path: str) -> Conversation:
        """Parse XLSX file using openpyxl."""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for XLSX files. Install with: pip install openpyxl")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Try to find a sheet with chat data
            sheet = None
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                if ws.max_row > 1:  # Has data
                    sheet = ws
                    break

            if sheet is None:
                raise ValueError("No sheets with data found")

            self.logger.info(f"Reading sheet: {sheet.title}")

            # Extract headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else '')

            # Detect column mapping
            column_map = self._detect_columns(headers)
            self.logger.info(f"Detected columns: {column_map}")

            # Get column indices
            timestamp_idx = headers.index(column_map['timestamp']) if 'timestamp' in column_map else None
            username_idx = headers.index(column_map['username'])
            message_idx = headers.index(column_map['message'])

            # Parse messages
            messages = []
            participants_set = set()

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Extract values
                    username = str(row[username_idx]).strip() if row[username_idx] else ''
                    content = str(row[message_idx]).strip() if row[message_idx] else ''

                    if not username or not content:
                        continue

                    # Parse timestamp
                    if timestamp_idx is not None:
                        timestamp_val = row[timestamp_idx]
                        if isinstance(timestamp_val, datetime):
                            timestamp = timestamp_val
                        elif timestamp_val:
                            timestamp = self._parse_timestamp(str(timestamp_val))
                        else:
                            from datetime import timedelta
                            timestamp = datetime(2024, 1, 1, 0, 0, 0) + timedelta(minutes=row_num)
                    else:
                        from datetime import timedelta
                        timestamp = datetime(2024, 1, 1, 0, 0, 0) + timedelta(minutes=row_num)

                    # Create message
                    msg = Message(
                        id=f"xlsx_{row_num}",
                        timestamp=timestamp,
                        sender=username,
                        content=content,
                        type=MessageType.TEXT,
                        platform='Excel',
                    )

                    messages.append(msg)
                    participants_set.add(username)

                except Exception as e:
                    self.logger.warning(f"Skipping row {row_num}: {e}")
                    continue

            workbook.close()

            if not messages:
                raise ValueError("No valid messages found in Excel file")

            # Create participants
            participants = [
                Participant(id=str(i), username=username)
                for i, username in enumerate(sorted(participants_set))
            ]

            # Create conversation
            conversation = Conversation(
                id=Path(file_path).stem,
                title=Path(file_path).stem.replace('_', ' ').title(),
                messages=messages,
                participants=participants,
                platform='Excel',
                conversation_type='unknown',
            )

            conversation.sort_messages()

            self.logger.info(f"Parsed {len(messages)} messages from {len(participants)} participants")

            return conversation

        except Exception as e:
            self.logger.error(f"XLSX parsing failed: {e}")
            raise

    def _parse_xls(self, file_path: str) -> Conversation:
        """Parse XLS file using xlrd."""
        if not HAS_XLRD:
            raise ImportError("xlrd is required for XLS files. Install with: pip install xlrd")

        try:
            workbook = xlrd.open_workbook(file_path)

            # Try to find a sheet with chat data
            sheet = None
            for sheet_idx in range(workbook.nsheets):
                ws = workbook.sheet_by_index(sheet_idx)
                if ws.nrows > 1:
                    sheet = ws
                    break

            if sheet is None:
                raise ValueError("No sheets with data found")

            self.logger.info(f"Reading sheet: {sheet.name}")

            # Extract headers
            headers = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]

            # Detect column mapping
            column_map = self._detect_columns(headers)
            self.logger.info(f"Detected columns: {column_map}")

            # Get column indices
            timestamp_idx = headers.index(column_map['timestamp']) if 'timestamp' in column_map else None
            username_idx = headers.index(column_map['username'])
            message_idx = headers.index(column_map['message'])

            # Parse messages
            messages = []
            participants_set = set()

            for row_num in range(1, sheet.nrows):
                try:
                    username = str(sheet.cell_value(row_num, username_idx)).strip()
                    content = str(sheet.cell_value(row_num, message_idx)).strip()

                    if not username or not content:
                        continue

                    # Parse timestamp
                    if timestamp_idx is not None:
                        timestamp_val = sheet.cell_value(row_num, timestamp_idx)
                        if isinstance(timestamp_val, float):
                            # Excel date format
                            timestamp = xlrd.xldate_as_datetime(timestamp_val, workbook.datemode)
                        elif timestamp_val:
                            timestamp = self._parse_timestamp(str(timestamp_val))
                        else:
                            from datetime import timedelta
                            timestamp = datetime(2024, 1, 1, 0, 0, 0) + timedelta(minutes=row_num)
                    else:
                        from datetime import timedelta
                        timestamp = datetime(2024, 1, 1, 0, 0, 0) + timedelta(minutes=row_num)

                    # Create message
                    msg = Message(
                        id=f"xls_{row_num}",
                        timestamp=timestamp,
                        sender=username,
                        content=content,
                        type=MessageType.TEXT,
                        platform='Excel',
                    )

                    messages.append(msg)
                    participants_set.add(username)

                except Exception as e:
                    self.logger.warning(f"Skipping row {row_num}: {e}")
                    continue

            if not messages:
                raise ValueError("No valid messages found in Excel file")

            # Create participants
            participants = [
                Participant(id=str(i), username=username)
                for i, username in enumerate(sorted(participants_set))
            ]

            # Create conversation
            conversation = Conversation(
                id=Path(file_path).stem,
                title=Path(file_path).stem.replace('_', ' ').title(),
                messages=messages,
                participants=participants,
                platform='Excel',
                conversation_type='unknown',
            )

            conversation.sort_messages()

            self.logger.info(f"Parsed {len(messages)} messages from {len(participants)} participants")

            return conversation

        except Exception as e:
            self.logger.error(f"XLS parsing failed: {e}")
            raise

    def _detect_columns(self, headers: List[str]) -> Dict[str, str]:
        """
        Detect which columns contain timestamp, username, and message.

        Args:
            headers: Excel column headers

        Returns:
            Dict mapping 'timestamp', 'username', 'message' to actual column names

        Raises:
            ValueError: If required columns cannot be detected
        """
        headers_lower = [h.lower().strip() for h in headers]
        mapping = {}

        # Find timestamp column
        for col in self.TIMESTAMP_COLUMNS:
            if col in headers_lower:
                mapping['timestamp'] = headers[headers_lower.index(col)]
                break

        # Find username column
        for col in self.USERNAME_COLUMNS:
            if col in headers_lower:
                mapping['username'] = headers[headers_lower.index(col)]
                break

        # Find message column
        for col in self.MESSAGE_COLUMNS:
            if col in headers_lower:
                mapping['message'] = headers[headers_lower.index(col)]
                break

        # Validate required columns found
        if 'username' not in mapping:
            raise ValueError(f"Could not find username column in: {headers}")

        if 'message' not in mapping:
            raise ValueError(f"Could not find message column in: {headers}")

        # Timestamp is optional
        if 'timestamp' not in mapping:
            self.logger.warning("No timestamp column found, using sequential timestamps")

        return mapping

    def _parse_timestamp(self, timestamp_str: str) -> datetime:
        """Parse timestamp string with various format support."""
        # Common timestamp formats
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y/%m/%d %H:%M:%S',
            '%Y/%m/%d %H:%M',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y %H:%M',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%dT%H:%M:%SZ',
        ]

        for fmt in formats:
            try:
                return datetime.strptime(timestamp_str, fmt)
            except ValueError:
                continue

        # Try ISO format
        try:
            return datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
        except:
            pass

        raise ValueError(f"Could not parse timestamp: {timestamp_str}")
